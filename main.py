"""
Karadel calc worker
====================
Небольшой Python-сервис, который снимает с Lovable тяжёлый расчёт.

Поток:
  Lovable -> POST /process {order_id}   (отвечаем 202 за доли секунды)
  фоновая задача: скачать чертёж и шаблон из Supabase -> Claude по чертежу
                  -> заполнить xlsx правкой XML в zip -> залить результат
                  -> статус заказа = "done"

Ничего не загружает целиком через ExcelJS/openpyxl: правятся ТОЛЬКО
два листа (sheet1.xml = Расчет, sheet6.xml = Сводная), остальные ~774
части копируются байт-в-байт. Выпадающие списки и условное
форматирование (extLst/x14) при этом сохраняются.

Запуск локально:  uvicorn main:app --reload
Деплой:           Railway (см. Procfile и README.md)
"""

import os
import io
import re
import json
import time
import base64
import logging
import zipfile
import traceback
from datetime import datetime, timezone

# Метка версии — видно по /health. Если после деплоя /health не показывает
# этот номер, значит на сервере СТАРЫЙ файл (загрузка не доехала).
WORKER_VERSION = "v25-qty-number-match-2026-06-25"
from xml.sax.saxutils import escape

import fitz  # PyMuPDF
from PIL import Image
import anthropic
from supabase import create_client
from fastapi import FastAPI, BackgroundTasks, Header, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger("worker")

# ---------------------------------------------------------------------------
# КОНФИГ — подгоните под свою схему Supabase, если имена отличаются
# ---------------------------------------------------------------------------
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
WORKER_SECRET = os.environ.get("WORKER_SECRET", "")
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-haiku-4-5-20251001")

# таблица заказов и её колонки
ORDERS_TABLE = "orders"
ID_COL = "id"
OVERRIDES_TABLE = "order_overrides"   # ручные цены менеджера по заявке
CATALOG_TABLE = "catalog_overrides"   # глобальные цены справочника (на все заявки)
STATUS_COL = "status"
LOG_COL = "processing_log"          # JSON-массив
DRAWING_PATH_COL = "drawing_url"    # путь файла чертежа внутри бакета DRAWINGS_BUCKET
REQUEST_PATH_COL = "request_url"    # путь файла заявки (JPG/PDF) внутри бакета REQUESTS_BUCKET
REQUEST_TEXT_COL = "request_text"   # текст заявки, вставленный менеджером прямо в форму
RESULT_PATH_COL = "result_url"      # сюда запишем путь готового файла
STATUS_DONE = "completed"           # статус при успехе (у вас "completed", не "done")
ERROR_MSG_COL = "error_message"     # сюда пишем текст ошибки

# бакеты хранилища
DRAWINGS_BUCKET = "drawings"
TEMPLATES_BUCKET = "templates"
RESULTS_BUCKET = "results"
PRICELIST_BUCKET = "pricelists"     # сюда оболочка кладёт прайс металлобазы «металл»
REQUESTS_BUCKET = "requests"        # сюда оболочка кладёт файл заявки (JPG/PDF)
# имя файла прайса «металл» в бакете PRICELIST_BUCKET (можно переопределить env-переменной)
PRICELIST_PATH = os.environ.get("PRICELIST_PATH", "metall.xlsx")

# путь активного шаблона в бакете TEMPLATES_BUCKET (задаётся env-переменной)
TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "")

MAX_PAGES = int(os.environ.get("MAX_PAGES", "15"))
# Порог «большого формата» в пунктах (1700 ≈ больше А2). Крупнее — сразу тяжёлый
# рендер с тайлами; мельче — сначала лёгкий проход, тяжёлый только если не прочиталось.
LARGE_FMT_PT = float(os.environ.get("LARGE_FMT_PT", "1700"))
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ВАЖНО: вставьте сюда ваш рабочий промпт распознавания чертежа из Lovable.
# Главное требование — Claude должен вернуть ТОЛЬКО JSON указанного вида.
EXTRACTION_PROMPT = """
Ты — инженер-расчётчик металлоконструкций. На изображениях — листы чертежа КМ.

СНАЧАЛА ОПРЕДЕЛИ ТИП ТАБЛИЦЫ НА ЧЕРТЕЖЕ (это критично):

ТИП 1 — «Техническая спецификация металла» (сводная по проекту/марке КМ):
  левый столбец — ТИП профиля (Швеллеры, Прокат листовой, Прокат угловой,
  Квадратные трубы, Двутавр…), есть столбцы масс по элементам конструкций и
  «Общая масса», внизу «Итого масса металла». Применяй РАЗДЕЛ «ТИП 1» ниже.

ТИП 2 — «Спецификация на отправочный элемент» (деталировка одного изделия,
  напр. «Траверса Б2»): столбцы «Марка/Поз.», «Кол-во шт», «Сечение», «Длина мм»,
  «Масса кг» с подстолбцами (т/н/шт/«элем.»/«общ.»), «Марка стали».
  Здесь ОДИН сортамент встречается в НЕСКОЛЬКИХ позициях разной длины.
  Применяй РАЗДЕЛ «ТИП 2» ниже.

Если есть обе — бери ту, что подробнее описывает изделие на этих листах.

ТИП 4 — «Деталь-пластина» (эскиз БЕЗ таблицы спецификации): нарисована ОДНА
  деталь — пластина/лист прямоугольной формы с габаритными размерами
  (длина, ширина) и толщиной S (напр. «S=20мм»), материал подписан (напр. 09Г2С),
  есть отверстия с подписью «⌀<d>, <N> отв». Таблицы сортамента НЕТ.
  Применяй РАЗДЕЛ «ТИП 4» ниже.

============================ РАЗДЕЛ «ТИП 4» ============================
(одиночная деталь-пластина из листа, эскиз без таблицы)

ЦЕЛЬ: снять РАЗМЕРЫ и материал. Массу НЕ считай сам — её посчитает программа.
Верни объект "plate" (а "positions" оставь пустым []):
- length_mm — бОльший габарит пластины, мм (напр. 2850)
- width_mm  — меньший габарит, мм (напр. 290)
- thickness_mm — толщина S из подписи «S=…мм», мм (напр. 20)
- grade — марка стали из подписи (напр. "09Г2С")
- name — наименование детали из заголовка (напр. "Пластина боковая Гит 32")
- holes — массив отверстий. Количество бери ИЗ ПОДПИСИ «<N> отв» (НЕ считай
  нарисованные кружки — подпись авторитетнее). Диаметр — из «⌀<d>».
  Пример: [{"d":18,"count":10}].
Если деталей-пластин несколько — верни массив "plates" из таких объектов.
Размеры читай ВНИМАТЕЛЬНО (эскиз рукописный); если цифра не читается — uncertain:true.

============================ РАЗДЕЛ «ТИП 2» ============================
(деталировочная спецификация отправочного элемента)

ЦЕЛЬ: выписать КАЖДУЮ строку спецификации ОТДЕЛЬНО (НЕ суммируй сам — сумму
по сортаменту посчитает программа). Твоя задача — точно прочитать таблицу.

ПРАВИЛА ТИП 2:
1) Пройди позиции ПО ПОРЯДКУ (поз. 1, 2, 3, … до последней) и для КАЖДОЙ верни
   отдельный объект в "positions". НИЧЕГО не объединяй и не складывай.
   Каждая строка таблицы спецификации = один объект.
2) Для каждой позиции возьми МАССУ из столбца «общ.» (общая масса позиции —
   там уже учтено «Кол-во шт»). Это поле mass_kg. НЕ бери «элем.» (масса 1 детали).
3) sortament — название по «Сводной»:
   - уголок «{70х6», «∟70×6», «L70x6» → "Уголок 70x6"
   - ЛИСТ/ПОЛОСА «−16×100», «-16x100» (первое число — ТОЛЩИНА): бери ТОЛЬКО
     толщину → "Лист 16". Ширину/длину в название НЕ бери.
   - КРУГ: знак диаметра перед числом = круг. Знак может быть «⌀»,«Ø»,«∅»,«Ç»,
     «q»,«d»,«d=». Пример: «Ç20»→"Круг 20", «Ç16»→"Круг 16". ВСЕГДА "Круг <N>".
     ЗАПРЕЩЕНО писать «Катанка» или «Труба» вместо круга.
   - труба — только если явно слово «труба».
4) НЕ ПРОПУСКАЙ ни одной строки, включая мелкие (0,1–0,6 кг) и последние в таблице.
   Сколько позиций в таблице (1..N) — столько объектов в "positions".
5) Добавь поле "pos" — номер позиции из спецификации (для самопроверки полноты).
6) СВАРКА: НЕ добавляй позицию «Сварка», НЕ включай швы в массу. Если есть строка
   «На сварку N%» — верни weld_pct=N (справочно). Иначе weld_pct=null.
7) НЕ включай крепёж из «Ведомости монтажных метизов» (болты/гайки/шайбы).
8) control_total_kg = масса элемента КАК В ЧЕРТЕЖЕ (со сваркой, напр. 108.96).
   НЕ вычитай сварку сам — это сделает программа.
9) "mark" — марка изделия из штампа (напр. "Траверса Б2").

Пример positions для ТИП 2 (каждая поз. отдельно, программа потом сложит):
[{"pos":"1","sortament":"Уголок 70x6","mass_kg":45.18,"grade":"С245"},
 {"pos":"2","sortament":"Уголок 63x5","mass_kg":18.90,"grade":"С245"},
 {"pos":"3","sortament":"Уголок 63x5","mass_kg":9.24,"grade":"С245"}, ...]

============================ РАЗДЕЛ «ТИП 1» ============================
Главный источник данных — таблица «Техническая спецификация металла»
(обычно лист 1–2). Извлеки из неё перечень стального проката.

КАК УСТРОЕНА ТАБЛИЦА:
- Крайний ЛЕВЫЙ столбец — ТИП профиля (напр. «Швеллеры по ГОСТ 8240-97»,
  «Прокат листовой горячекатаный», «Прокат угловой равнополочный»,
  «Квадратные трубы», «Двутавр по ГОСТ 26020-83»). Он РАСТЯНУТ на несколько строк
  и относится КО ВСЕМ строкам-размерам под ним, до следующего типа.
- Столбец «Наименование или марка металла» — марка стали (С255, С345 и т.п.).
- Столбец «Номер или размер профиля в мм» — конкретный размер: [8, [10, [12;
  t5, t8, t10, t14…; L45x5, L90x6, L100x10; []60x5, []100x5; 30К1, 23К1, 30Б1.
- Средние столбцы «Масса металла по элементам конструкций» (Колонна, Ферма,
  Прогоны, связи, Раскос, Стеновые прогоны) — НЕ бери их по отдельности.
- Крайний ПРАВЫЙ столбец «Общая масса» — ИТОГОВАЯ масса этого размера.
  БЕРИ МАССУ ТОЛЬКО ОТСЮДА.

ЕДИНИЦЫ МАССЫ — ОЧЕНЬ ВАЖНО:
- В шапке столбца массы всегда подписана единица: «Общая масса, кг» ИЛИ
  «Общая масса, т» (тонны). Сначала ОПРЕДЕЛИ единицу по шапке.
- В ответе mass_kg и control_total_kg ВСЕГДА указывай В КИЛОГРАММАХ.
- Если шапка в тоннах («т», «тн», «тонн») — УМНОЖЬ каждое значение НА 1000
  (45,04 т → 45040). Если в кг — оставь как есть.
- Десятичная запятая = точка (45,04 → 45.04), затем перевод в кг.
- Подстраховка: итоговая масса металлоконструкции обычно ТЫСЯЧИ кг.
  Если после перевода сумма вышла в десятках (напр. 45) — ты забыл умножить
  тонны на 1000, перечитай шапку единиц.

ЧТО ДЕЛАТЬ (по каждой строке-размеру):
1) Если в столбце «Общая масса» стоит число — это позиция. Возьми её (в кг).
2) Если «Общая масса» пустая (нет массы) — ПРОПУСТИ строку (не вноси).
3) Тип бери из левого растянутого столбца, размер — из столбца размера,
   массу — из «Общая масса» (переведённую в кг), марку — из столбца марки.

ЧТО ПРОПУСКАТЬ ВСЕГДА:
- Строки «Всего профиля» (это подытоги — НЕ нужны).
- Строку «Итого масса металла» (общий итог — НЕ нужна).
- Блок «в том числе по маркам» (С235 / С255 / С345 — НЕ нужен).
- Строки без массы в столбце «Общая масса».

ПЕРЕВОД ТИПА В НАЗВАНИЕ (как в нашей базе «Сводная»); это база синонимов,
сопоставляй гибко:
- «Швеллеры …», «Швеллер», «[N»            → "Швеллер <номер>"  (напр. [10 → "Швеллер 10")
- «Прокат листовой …», «Лист», «t<N>»       → "Лист t<толщина>"  (t10 → "Лист t10")
- «Прокат угловой равнополочный …», «Уголок», «L…» → "Уголок <размер>" (L100x10 → "Уголок 100x10")
- ТРУБЫ. Круглая или профильная — определяй ТОЛЬКО по ЛЕВОМУ столбцу (вид
  профиля/ГОСТ), а НЕ по количеству чисел в размере. Это критично: и круглая,
  и квадратная труба могут писаться двумя числами (273x4 и 250x6 выглядят
  одинаково, но это РАЗНЫЕ трубы).
  • КВАДРАТНАЯ / ПРЯМОУГОЛЬНАЯ (профильная): левый столбец «Трубы квадратные…»,
    «Трубы прямоугольные…», ГОСТ 30245 / 8639 / 8645, слова «профильная»/«[]…».
    → "Труба проф. <размер>". Квадратная может быть записана как «250x6»
    (это 250×250×6) или «250×250×6»; прямоугольная — тремя числами «160×120×4».
  • КРУГЛАЯ: левый столбец «электросварные»/«бесшовные»/«водогазопроводные/ВГП»,
    ГОСТ 10704 / 10705 / 8732 / 8734 / 3262, символы «Ф»/«Ø», размер диаметр×толщина
    (273x4, Ф159х4). → "Труба <диаметр>x<толщина>" (273x4 → "Труба 273x4").
    НЕ пиши такие как «Труба проф.».
- «Двутавр …», «Балка», «<N>К<n>», «<N>Б<n>»→ "Балка <обозначение>" (30К1 → "Балка 30К1", 30Б1 → "Балка 30Б1")
ВАЖНО: обозначения вида «30К1», «23К1», «30Б1» — это ДВУТАВРЫ (балки),
а НЕ листы. Никогда не пиши их как «Лист».

ПОЛЯ ПОЗИЦИИ:
- sortament: переведённое название (как в «Сводной»).
- grade: марка стали.
- mass_kg: число из «Общая масса», ПЕРЕВЕДЁННОЕ В КИЛОГРАММЫ (обязательно).
- qty: если есть, иначе пропусти.
- uncertain: true, ЕСЛИ ты не уверен в этой позиции (чертёж размыт, цифры/название
  плохо читаются, неоднозначно). НЕ ФАНТАЗИРУЙ — лучше поставь uncertain:true.
  Если всё чётко — uncertain:false или не указывай.

ЗАПРЕЩЕНО: выдумывать профили/размеры/массы, использовать римские цифры,
повторять один и тот же размер дважды, включать позиции без массы.
Если число не разобрать — поставь uncertain:true (НЕ угадывай значение).

Самопроверка: сумма всех mass_kg (в кг) должна примерно совпасть с «Итого масса
металла» из таблицы (тоже переведённой в кг). Если расходится в ~1000 раз —
ты перепутал тонны и килограммы, перечитай шапку единиц столбца массы.
Дополнительно верни это контрольное число в поле "control_total_kg"
(значение строки «Итого масса металла» В КИЛОГРАММАХ; если в шапке тонны —
умножь на 1000; если строки нет — null).

Верни СТРОГО JSON без пояснений и markdown:
{
  "drawing": "обозначение из штампа, если видно",
  "mark": "марка изделия для ТИП 2 (напр. «Траверса Б2»); для ТИП 1 — null",
  "weld_pct": null,
  "control_total_kg": 70103.84,
  "plate": null,
  "positions": [
    {"sortament": "Швеллер 10", "grade": "С255", "mass_kg": 5803.68, "qty": 1},
    {"sortament": "Лист t10", "grade": "С345", "mass_kg": 8021.71, "qty": 1},
    {"sortament": "Балка 30К1", "grade": "С345", "mass_kg": 15152, "qty": 1}
  ]
}
Только JSON, ничего больше.
""".strip()

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
app = FastAPI(title="Karadel calc worker")


# ---------------------------------------------------------------------------
# Вспомогательное
# ---------------------------------------------------------------------------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


def col_to_idx(letters):
    n = 0
    for c in letters:
        n = n * 26 + (ord(c) - 64)
    return n


def idx_to_col(idx):
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


# стальные сортаменты живут в колонках N..BA (40 штук)
STEEL_COLUMNS = [idx_to_col(i) for i in range(col_to_idx("N"), col_to_idx("BA") + 1)]


def fetch_order(order_id):
    res = sb.table(ORDERS_TABLE).select("*").eq(ID_COL, order_id).single().execute()
    return res.data


def db_update(order_id, fields):
    sb.table(ORDERS_TABLE).update(fields).eq(ID_COL, order_id).execute()


def load_overrides(order_id):
    """Ручные цены менеджера для заявки: {имя_материала: цена_за_кг}."""
    out = {}
    try:
        res = sb.table(OVERRIDES_TABLE).select("match_name, price").eq("order_id", order_id).execute()
        for row in (res.data or []):
            nm = (row.get("match_name") or "").strip()
            pr = row.get("price")
            if nm and pr is not None:
                out[nm] = float(pr)
    except Exception as e:
        log.error("[load_overrides] не удалось прочитать ручные цены: %s", e)
    return out


def load_catalog_overrides():
    """Глобальные цены справочника (на все заявки): {имя_материала: цена_за_кг}.
    Ошибку НЕ глушим — пусть всплывёт в лог обработки (CATALOG_ERROR)."""
    out = {}
    res = sb.table(CATALOG_TABLE).select("match_name, price").execute()
    for row in (res.data or []):
        nm = (row.get("match_name") or "").strip()
        pr = row.get("price")
        if nm and pr is not None:
            out[nm] = float(pr)
    return out


def log_run(*, drawing_name, check_status,
            positions_count=None, sum_mass=None, expected_total=None,
            review=None, cost_cents=None, request_id=None, log_text=None):
    """Пишет одну строку-журнал в worker_runs по итогам обработки чертежа.
    Никогда не роняет основную обработку: ошибку журнала просто логируем."""
    review = review or {}
    payload = {
        "drawing_name":   drawing_name,
        "request_id":     request_id,
        "worker_version": WORKER_VERSION,
        "model":          CLAUDE_MODEL,
        "check_status":   check_status,        # "OK" / "MISMATCH" / "NO_CONTROL"
        "positions_count": positions_count,
        "sum_mass":       sum_mass,
        "expected_total": expected_total,
        "review_total":   review.get("total"),
        "review_yellow":  review.get("yellow"),
        "review_orange":  review.get("orange"),
        "review_red":     review.get("red"),
        "cost_cents":     cost_cents,
        "log_text":       log_text,
    }
    try:
        sb.table("worker_runs").insert(payload).execute()
    except Exception as e:
        log.error("[log_run] не удалось записать журнал прогона: %s", e)


def get_template_bytes():
    if not TEMPLATE_PATH:
        raise RuntimeError("TEMPLATE_PATH не задан (путь активного шаблона в бакете templates)")
    return sb.storage.from_(TEMPLATES_BUCKET).download(TEMPLATE_PATH)


def load_metall_pricelist():
    """Читает загруженный прайс металлобазы «металл» из бакета PRICELIST_BUCKET.
    Возвращает {имя_материала: цена_за_тонну}. Имя — колонка B (с 16-й строки),
    цена — колонка H («Цена за 1 тн»). Цены здесь УЖЕ за тонну (как в «Сводной»).
    Если файла нет или нет openpyxl — возвращает пусто, расчёт продолжается."""
    out = {}
    try:
        data = sb.storage.from_(PRICELIST_BUCKET).download(PRICELIST_PATH)
    except Exception as e:
        log.info("[pricelist] прайс «металл» не загружен (%s) — пропускаю", e)
        return out
    try:
        import openpyxl
    except Exception:
        log.error("[pricelist] нет openpyxl — добавьте 'openpyxl' в requirements.txt")
        return out
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb["металл"] if "металл" in wb.sheetnames else wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=16, values_only=True):
            if len(row) < 8:
                continue
            name = row[1]   # B
            price = row[7]  # H
            if name is None:
                continue
            name = str(name).strip()
            try:
                price = float(price)
            except (TypeError, ValueError):
                continue
            if name and price > 0:
                out[name] = price
    except Exception as e:
        log.error("[pricelist] ошибка чтения прайса: %s", e)
    return out


def download_drawing(order):
    path = order.get(DRAWING_PATH_COL)
    if not path:
        raise RuntimeError(f"в заказе нет {DRAWING_PATH_COL}")
    return sb.storage.from_(DRAWINGS_BUCKET).download(path)


def download_drawing_by_path(path):
    return sb.storage.from_(DRAWINGS_BUCKET).download(path)


def load_order_drawings(order_id):
    """Список чертежей заказа из order_drawings (отсортирован по position_index).
    Если таблица пустая для заказа — вернёт []."""
    try:
        res = (sb.table("order_drawings")
               .select("*").eq("order_id", order_id)
               .order("position_index").execute())
        return res.data or []
    except Exception as e:
        log.error("[multi] не удалось прочитать order_drawings: %s", e)
        return []


def update_order_drawing(row_id, fields):
    try:
        sb.table("order_drawings").update(fields).eq("id", row_id).execute()
    except Exception as e:
        log.error("[multi] не удалось обновить order_drawings %s: %s", row_id, e)


def upload_result(path, data):
    # upsert=true чтобы перезапись при повторном расчёте не падала
    sb.storage.from_(RESULTS_BUCKET).upload(
        path, data, {"content-type": XLSX_MIME, "upsert": "true"}
    )


# ---------------------------------------------------------------------------
# Чертёж -> картинки для Claude
# ---------------------------------------------------------------------------
def drawing_to_images(data, filename):
    raw_images = []
    is_pdf = filename.lower().endswith(".pdf") or data[:4] == b"%PDF"
    if is_pdf:
        doc = fitz.open(stream=data, filetype="pdf")
        for i, page in enumerate(doc):
            if i >= MAX_PAGES:
                break
            long_side = max(page.rect.width, page.rect.height) or 1000
            zoom = min(3.0, 1500.0 / long_side)
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
            raw_images.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
    else:
        raw_images.append(Image.open(io.BytesIO(data)).convert("RGB"))

    out = []
    for img in raw_images:
        img.thumbnail((1500, 1500))
        buf = io.BytesIO()
        img.save(buf, "JPEG", quality=75)
        out.append(base64.b64encode(buf.getvalue()).decode())
    return out


# ---------------------------------------------------------------------------
# Claude -> позиции
# ---------------------------------------------------------------------------
def parse_json(text):
    t = text.strip()
    t = re.sub(r"^```(?:json)?", "", t).strip()
    t = re.sub(r"```$", "", t).strip()
    start, end = t.find("{"), t.rfind("}")
    if start != -1 and end != -1:
        t = t[start:end + 1]
    return json.loads(t)


def _norm_sort_key(s):
    """Ключ сортамента для группировки: без регистра/пробелов, x/х/* унифицированы."""
    s = str(s or "").lower().replace("ё", "е")
    s = s.replace("х", "x").replace("*", "x").replace("×", "x")
    return re.sub(r"\s+", "", s)


def _uniq_sortaments(positions):
    return {_norm_sort_key(p.get("sortament")) for p in positions if p.get("sortament")}


def _aggregate_by_sortament(positions):
    """Складывает массы позиций с одинаковым сортаментом в одну строку.
    Берёт первое читаемое название сортамента и grade. Количество (qty) — 1
    (масса уже общая по сортаменту на изделие)."""
    groups = {}
    order = []
    for p in positions:
        srt = p.get("sortament")
        if not srt:
            continue
        key = _norm_sort_key(srt)
        try:
            m = float(p.get("mass_kg") or 0)
        except (TypeError, ValueError):
            m = 0.0
        if key not in groups:
            groups[key] = {"sortament": srt, "grade": p.get("grade"),
                           "mass_kg": 0.0, "qty": 1}
            order.append(key)
        groups[key]["mass_kg"] += m
        if not groups[key].get("grade") and p.get("grade"):
            groups[key]["grade"] = p.get("grade")
    out = []
    for key in order:
        g = groups[key]
        g["mass_kg"] = round(g["mass_kg"], 2)
        out.append(g)
    return out


def extract_positions(images_b64, target_sheets=None, preselected=False):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    content = [
        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b}}
        for b in images_b64
    ]
    prompt = EXTRACTION_PROMPT
    if target_sheets:
        sheets_str = ", ".join(str(s) for s in target_sheets)
        pref = SHEET_READ_PREFIX if preselected else SHEET_FILTER_PREFIX
        # .replace, а НЕ .format: в подсказке есть пример JSON с фигурными скобками
        prompt = pref.replace("{sheets}", sheets_str) + EXTRACTION_PROMPT
    content.append({"type": "text", "text": prompt})
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    return parse_json(text)


# Общие правила чтения таблицы спецификации листа (для обоих режимов).
TABLE_READ_RULES = """КАК ЧИТАТЬ ТАБЛИЦУ СПЕЦИФИКАЦИИ (столбцы:
Поз. | Обозначение | Наименование | Кол. | Масса ед., кг | Примечание):
- МАССА: в столбце стоит «Масса ед.» — это масса ОДНОЙ детали. Массу позиции
  считай как mass_kg = «Масса ед.» × «Кол.». (Пример: Кол.=2, Масса ед.=83,2 → 166,4.)
- На листе может быть НЕСКОЛЬКО изделий, разделённых строками-заголовками
  (напр. «Площадка ПЛ1», «Накладка Н1», «Короб К1»). Выпиши позиции ВСЕХ изделий
  этого листа. НЕ пропускай ни одной строки (их может быть 20+).
- Для КАЖДОЙ позиции добавь поле "izdelie" — название изделия (заголовок раздела),
  под которым стоит строка (напр. "Площадка ПЛ1", "Накладка Н1", "Короб К1").
- Верни также массив "izdelia" — заявленные массы изделий из строк-заголовков
  (для контроля), напр. "izdelia":[{"name":"Площадка ПЛ1","mass_kg":1549.3},
  {"name":"Накладка Н1","mass_kg":90.15},{"name":"Короб К1","mass_kg":123.44}].
- НЕ включай строку «Наплавленный металл» (это сварка, не сортамент).
- НЕ включай строки-заголовки изделий и строку «Итого».
- В столбце «Обозначение» может стоять ссылка на ДРУГОЙ чертёж (напр. «43-2023-ЭВ2,
  лист 19») — это нормально, ИГНОРИРУЙ её и всё равно читай таблицу.
- Сортамент (sortament) переводи как в «Сводной»:
  • «Швеллер 12», «[12» → "Швеллер 12";  «Швеллер 8» → "Швеллер 8"
  • «Уголок 90х7», «L90x7» → "Уголок 90x7";  «Уголок 80х6» → "Уголок 80x6";
    «Уголок 40х4» → "Уголок 40x4"
  • «Сталь круглая ⌀16», «ф16», «d16» → "Круг 16";  «ф6» → "Круг 6"
  • листовая сталь «Сталь 660х3 … лист», «660x3» (толщина — наименьший размер, тут 3)
    → "Лист 3";  «900х3 … лист» → "Лист 3"
  • «Сетка №35 …» → "Сетка 35"
"""

# Режим А: страница листа УЖЕ отобрана (даны её фрагменты-тайлы). Штамп искать
# не нужно — просто прочитать таблицу целиком.
SHEET_READ_PREFIX = """ВАЖНО: эти изображения — фрагменты ОДНОГО листа № {sheets},
он уже отобран по штампу. Тебе НЕ нужно искать штамп или другие листы.
Просто прочитай таблицу спецификации металла этого листа ПОЛНОСТЬЮ.

""" + TABLE_READ_RULES + """
ДОПОЛНИТЕЛЬНО верни "sheets_seen": ["{sheets}"].
=====================================================================

"""

# Режим Б (запасной): страница не отобрана — даны все страницы, надо найти лист
# по ШТАМПУ самому.
SHEET_FILTER_PREFIX = """ОГРАНИЧЕНИЕ ПО ЛИСТАМ — ВЫПОЛНИ В ПЕРВУЮ ОЧЕРЕДЬ.
Тебя просят рассчитать ТОЛЬКО лист(ы): {sheets}.
Номер листа написан в ШТАМПЕ чертежа — рамка справа внизу, графа «Лист»
(между «Стадия» и «Листов»). Это НЕ номер страницы и НЕ «лист NN» из столбца
«Обозначение». Сверяйся ТОЛЬКО со штампом.
ПОРЯДОК:
1) Найди страницу(ы), где в штампе стоит один из номеров: {sheets}.
2) Читай спецификацию ТОЛЬКО с них. Если номера нет в штампах — "positions": [].

""" + TABLE_READ_RULES + """
ДОПОЛНИТЕЛЬНО верни "sheets_seen" — все номера листов из штампов страниц.
=====================================================================

"""


def _parse_sheet_numbers(text):
    """Из текста заявки достаёт номера листов, если менеджер написал команду вида
    «посчитай лист 28, 30», «листы 28 30», «лист №30», «л. 30».
    Возвращает список строк-номеров в порядке появления, без повторов. Иначе []."""
    if not text:
        return []
    t = str(text).lower().replace("ё", "е")
    sheets = []
    # «лист(ы/а) [№] 28, 30 [и] 31» — забираем все числа из хвоста после слова «лист»
    for m in re.finditer(r"\bлист(?:ы|а)?\s*№?\s*([\d\s,;]+(?:\s*и\s*\d+)*)", t):
        for num in re.findall(r"\d+", m.group(1)):
            if num not in sheets:
                sheets.append(num)
    # запасной вариант «л. 30»
    if not sheets:
        for m in re.finditer(r"\bл\.?\s*№?\s*(\d+)", t):
            num = m.group(1)
            if num not in sheets:
                sheets.append(num)
    return sheets


REQUEST_PROMPT = """Ты разбираешь ЗАЯВКУ заказчика на металлоконструкции (фото/скан/PDF таблицы).
Извлеки КАЖДУЮ строку таблицы заявки. Верни СТРОГО JSON без пояснений и без ```:
{
  "items": [
    {"pos": "1", "mark": "ТМs-68", "name": "Траверса", "qty": 1359, "unit": "шт"},
    ...
  ]
}
Правила:
- "mark" — артикул/обозначение/марка изделия (например «ТМs-68», «ТМ73», «Т13.16», «Б2»).
  Бери его из колонки «Артикул»/«Обозначение» или из начала наименования. Сохраняй как написано.
- "name" — наименование без марки (например «Траверса», «Опора скользящая»). Если нет — пусто.
- "qty" — количество ЧИСЛОМ (убери пробелы-разделители тысяч: «1 359» -> 1359). Если нет — null.
- "unit" — единица как в заявке (шт, компл, м, кг, т). Если нет — пусто.
- Бери ВСЕ строки подряд, ничего не пропускай и не фильтруй (даже провод, изоляторы, зажимы).
- Не придумывай значения. Чего не видно — оставь пустым/null.
"""


def _request_xlsx_to_text(data):
    """Excel-заявка -> текст таблицы (все листы, по строкам). openpyxl уже в зависимостях."""
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"# Лист: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def _request_docx_to_text(data):
    """Word-заявка -> текст. Извлекает и абзацы, и ТАБЛИЦЫ (построчно, ячейки через |).
    Это важно: в заявках позиции обычно лежат в таблице Word."""
    import zipfile, io as _io
    z = zipfile.ZipFile(_io.BytesIO(data))
    xml = z.read("word/document.xml").decode("utf-8", "ignore")

    def _unescape(s):
        return (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
                 .replace("&quot;", '"').replace("&#39;", "'"))

    def _para_text(block):
        # текст одного абзаца: все <w:t>, табы/переводы строк как пробелы
        block = re.sub(r"<w:tab\b[^>]*/>", " ", block)
        block = re.sub(r"<w:br\b[^>]*/>", " ", block)
        return _unescape("".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", block, re.S))).strip()

    def _cell_text(tc):
        # текст ячейки = склейка её абзацев
        paras = re.findall(r"<w:p\b.*?</w:p>", tc, re.S)
        return " ".join(t for t in (_para_text(p) for p in paras) if t).strip()

    lines = []
    # идём по документу последовательно: и таблицы, и абзацы вне таблиц
    # 1) таблицы -> строки "ячейка | ячейка | ..."
    for tbl in re.findall(r"<w:tbl>.*?</w:tbl>", xml, re.S):
        for tr in re.findall(r"<w:tr\b.*?</w:tr>", tbl, re.S):
            cells = re.findall(r"<w:tc>.*?</w:tc>", tr, re.S)
            vals = [_cell_text(c) for c in cells]
            if any(vals):
                lines.append(" | ".join(vals))
    # 2) текст вне таблиц (заголовок/контекст): убираем таблицы, картинки, теги
    no_tbl = re.sub(r"<w:tbl>.*?</w:tbl>", "", xml, flags=re.S)
    no_tbl = re.sub(r"<w:drawing>.*?</w:drawing>", "", no_tbl, flags=re.S)
    no_tbl = re.sub(r"<w:tab\b[^>]*/>", " ", no_tbl)
    extra = _unescape(" ".join(re.findall(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", no_tbl, re.S)))
    extra = re.sub(r"<[^>]+>", " ", extra)          # на всякий случай выкидываем любые теги
    extra = re.sub(r"\s{2,}", " ", extra).strip()
    if extra:
        lines.append(extra)
    return "\n".join(lines)


PAGE_MAP_PROMPT = """Тебе даны страницы чертежа по порядку (стр. 1, 2, 3 …).
Для КАЖДОЙ страницы посмотри в ШТАМП — рамку справа внизу, графа «Лист»
(между «Стадия» и «Листов»). Верни номер листа из этой графы.
Верни СТРОГО JSON без пояснений:
{"pages":[{"page":1,"sheet":"5"},{"page":2,"sheet":"5.1"},{"page":3,"sheet":"30"}]}
Если на странице штампа/номера листа нет — "sheet": null. Только JSON."""


def map_pages_to_sheets(images_b64):
    """Один дешёвый проход: какой номер листа в штампе на каждой странице.
    Возвращает dict {номер_листа(str): индекс_страницы(0-based)}."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    content = [
        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b}}
        for b in images_b64
    ]
    content.append({"type": "text", "text": PAGE_MAP_PROMPT})
    resp = client.messages.create(
        model=CLAUDE_MODEL, max_tokens=1024,
        messages=[{"role": "user", "content": content}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    out = {}
    try:
        data = parse_json(text)
        for rec in data.get("pages", []):
            s = rec.get("sheet")
            p = rec.get("page")
            if s is None or p is None:
                continue
            key = str(s).strip().lower().replace("лист", "").strip()
            try:
                idx = int(p) - 1
            except (TypeError, ValueError):
                continue
            if key and key not in out and idx >= 0:
                out[key] = idx
    except Exception:
        return {}
    return out


def _page_long_side_pt(data, filename, page_index):
    """Длинная сторона страницы в пунктах (1 пт = 1/72 дюйма). Для не-PDF — None."""
    is_pdf = filename.lower().endswith(".pdf") or data[:4] == b"%PDF"
    if not is_pdf:
        return None
    try:
        doc = fitz.open(stream=data, filetype="pdf")
        r = doc[page_index].rect
        return max(r.width, r.height)
    except Exception:
        return None


def render_page_single(data, filename, page_index, cap=1500):
    """Лёгкий рендер ОДНОЙ страницы в одну картинку (как в обычном расчёте)."""
    is_pdf = filename.lower().endswith(".pdf") or data[:4] == b"%PDF"
    if not is_pdf:
        img = Image.open(io.BytesIO(data)).convert("RGB")
    else:
        doc = fitz.open(stream=data, filetype="pdf")
        page = doc[page_index]
        long_side = max(page.rect.width, page.rect.height) or 1000
        zoom = min(3.0, float(cap) / long_side)
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img.thumbnail((cap, cap))
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=78)
    return [base64.b64encode(buf.getvalue()).decode()]


def render_page_tiles(data, filename, page_index, max_tile=1500, overlap=140):
    """Рендерит ОДНУ страницу PDF в высоком разрешении и режет на тайлы ≤max_tile,
    чтобы плотная таблица была читаемой. Возвращает список base64-JPEG."""
    is_pdf = filename.lower().endswith(".pdf") or data[:4] == b"%PDF"
    if not is_pdf:
        # одиночная картинка — режем её саму
        img = Image.open(io.BytesIO(data)).convert("RGB")
    else:
        doc = fitz.open(stream=data, filetype="pdf")
        page = doc[page_index]
        long_side = max(page.rect.width, page.rect.height) or 1000
        zoom = min(6.0, 3600.0 / long_side)   # выше, чем общий рендер (там 1500)
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

    W, H = img.size
    if max(W, H) <= max_tile:
        tiles = [img]
    else:
        step = max_tile - overlap
        xs = list(range(0, max(W - overlap, 1), step)) or [0]
        ys = list(range(0, max(H - overlap, 1), step)) or [0]
        tiles = []
        for y in ys:
            for x in xs:
                tile = img.crop((x, y, min(x + max_tile, W), min(y + max_tile, H)))
                if tile.size[0] >= 60 and tile.size[1] >= 60:
                    tiles.append(tile)
    out = []
    for t in tiles[:12]:                       # предохранитель: не более 12 тайлов
        buf = io.BytesIO()
        t.save(buf, "JPEG", quality=82)
        out.append(base64.b64encode(buf.getvalue()).decode())
    return out


def extract_request_items_from_text(text):
    """Распознаёт заявку из ТЕКСТА (Word/Excel) -> {"items": [...]}. Тот же промпт."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    content = [{"type": "text", "text": REQUEST_PROMPT
                + "\n\nНиже ТЕКСТ заявки (из Word/Excel). Разбери его как таблицу заявки:\n\n"
                + text[:60000]}]
    resp = client.messages.create(
        model=CLAUDE_MODEL, max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )
    out = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    try:
        data = parse_json(out)
    except Exception:
        return {"items": []}
    return data if isinstance(data, dict) and "items" in data else {"items": []}


def read_request_items(data, filename):
    """Единая точка чтения заявки любого формата -> {"items":[...]}.
    Excel/Word — через текст (точнее), PDF/картинки — через распознавание изображения."""
    fn = (filename or "").lower()
    try:
        if fn.endswith((".xlsx", ".xlsm", ".xls")):
            return extract_request_items_from_text(_request_xlsx_to_text(data))
        if fn.endswith(".docx"):
            return extract_request_items_from_text(_request_docx_to_text(data))
        if fn.endswith((".txt", ".csv")):
            txt = data.decode("utf-8", "ignore") if isinstance(data, (bytes, bytearray)) else str(data)
            return extract_request_items_from_text(txt)
    except Exception as e:
        log.error("[request] не смог прочитать %s как текст: %s", filename, e)
        # упадём в распознавание картинкой ниже
    # PDF / JPG / PNG — старый путь через изображения
    imgs = drawing_to_images(data, filename)
    return extract_request_items(imgs)


def extract_request_items(images_b64):
    """Распознаёт заявку (фото/PDF) -> {"items": [{pos, mark, name, qty, unit}, ...]}."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    content = [
        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b}}
        for b in images_b64
    ]
    content.append({"type": "text", "text": REQUEST_PROMPT})
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    try:
        data = parse_json(text)
    except Exception:
        return {"items": []}
    return data if isinstance(data, dict) and "items" in data else {"items": []}


def download_request(order):
    """Скачивает файл заявки из бакета REQUESTS_BUCKET (если он прикреплён)."""
    path = order.get(REQUEST_PATH_COL)
    if not path:
        return None
    try:
        return sb.storage.from_(REQUESTS_BUCKET).download(path)
    except Exception as e:
        log.error("[request] не удалось скачать заявку: %s", e)
        return None


def _norm_mark(s):
    """Нормализует марку для сравнения: убирает дефисы/пробелы/точки, регистр, ё->е.
    «ТМ-1» = «ТМ1» = «ТМ 1»; «Т13.16» = «т1316»."""
    s = str(s or "").lower().replace("ё", "е")
    return re.sub(r"[\s\-_.,]+", "", s)


def _mark_cores(s):
    """Достаёт «ядра марок» из строки: буква(ы)+цифры, напр.
    «Траверс Б-2» -> {'б2'}; «Стойка ТС15» -> {'тс15'}; «ТМ-1, ТМ-2» -> {'тм1','тм2'}.
    Возвращает множество нормализованных ядер."""
    s = str(s or "").lower().replace("ё", "е")
    cores = set()
    # шаблоны вида: буквы + (необязат. дефис/пробел) + цифры  (Б2, ТС-15, ТМ 1, Т13.16)
    for m in re.findall(r"[a-zа-я]{1,4}[\s\-]?\d+(?:[.\-]\d+)?", s):
        cores.add(re.sub(r"[\s\-_.,]+", "", m))
    return cores


def match_request_to_drawing(request_items, drawing_mark):
    """Ищет в заявке количество для марки чертежа.
    Возвращает (qty, status, matched_item):
      status: "green" — надёжное совпадение; "yellow" — частичное; "red" — не найдено."""
    dn = _norm_mark(drawing_mark)
    d_cores = _mark_cores(drawing_mark)
    if not dn and not d_cores:
        return None, "red", None

    # 1) точное совпадение нормализованных марок целиком
    for it in request_items:
        if dn and _norm_mark(it.get("mark")) == dn:
            return it.get("qty"), "green", it

    # 2) совпадение по «ядру марки» (Б2 == Б-2, ТС15 == ТС-15) — надёжно -> green.
    #    Сравниваем ядра из марки чертежа с ядрами из марки И наименования заявки.
    if d_cores:
        core_hits = []
        for it in request_items:
            it_cores = _mark_cores(it.get("mark")) | _mark_cores(it.get("name"))
            if d_cores & it_cores:   # пересечение ядер
                core_hits.append(it)
        if len(core_hits) == 1:
            return core_hits[0].get("qty"), "green", core_hits[0]
        if len(core_hits) > 1:
            # несколько кандидатов с тем же ядром — не угадываем, отдаём менеджеру
            return None, "yellow", None

    # 2.5) запасной: по УНИКАЛЬНОМУ числу из >=2 цифр (напр. ГИТ«32»/«52»),
    #      когда буквы марки прочитались криво («Гчт»≈«ГИТ»). Берём, только если
    #      такое число есть у РОВНО одной строки заявки -> надёжно (green).
    d_nums = set(re.findall(r"\d{2,}", _norm_mark(drawing_mark)))
    if d_nums:
        num_hits = []
        for it in request_items:
            it_nums = set(re.findall(r"\d{2,}",
                          _norm_mark(it.get("mark")) + " " + _norm_mark(it.get("name"))))
            if d_nums & it_nums:
                num_hits.append(it)
        if len(num_hits) == 1:
            return num_hits[0].get("qty"), "green", num_hits[0]

    # 3) запасной вариант: вхождение нормализованных строк (>=3 симв.) -> yellow
    cand = []
    for it in request_items:
        mn = _norm_mark(it.get("mark"))
        if len(dn) >= 3 and len(mn) >= 3 and (dn in mn or mn in dn):
            cand.append(it)
    if len(cand) == 1:
        return cand[0].get("qty"), "yellow", cand[0]
    return None, "red", None


# ---------------------------------------------------------------------------
# Правка xlsx через zip (без ExcelJS / openpyxl-save)
# ---------------------------------------------------------------------------
def _build_cell(ref, attrs, value, kind, style_id=None):
    if style_id is not None:
        s = f' s="{style_id}"'
    else:
        m = re.search(r'\bs="(\d+)"', attrs or "")
        s = f' s="{m.group(1)}"' if m else ""
    if kind == "n":
        return f'<c r="{ref}"{s}><v>{value}</v></c>'
    text = escape(str(value))
    return f'<c r="{ref}"{s} t="inlineStr"><is><t xml:space="preserve">{text}</t></is></c>'


def set_cell(xml, ref, value, kind, style_id=None):
    col = re.match(r"[A-Z]+", ref).group()
    row = re.search(r"\d+", ref).group()
    col_idx = col_to_idx(col)

    # 1) ячейка уже есть -> заменить
    pat = re.compile(r'<c r="' + ref + r'"([^>]*?)(?:/>|>.*?</c>)', re.S)
    if pat.search(xml):
        return pat.sub(lambda m: _build_cell(ref, m.group(1), value, kind, style_id), xml, count=1)

    # 2) ячейки нет -> вставить в нужную строку в порядке колонок
    rowpat = re.compile(r'(<row r="' + row + r'"[^>]*>)(.*?)(</row>)', re.S)
    m = rowpat.search(xml)
    new_cell = _build_cell(ref, "", value, kind, style_id)
    if not m:
        return xml.replace("</sheetData>", f'<row r="{row}">{new_cell}</row></sheetData>', 1)
    head, body, tail = m.group(1), m.group(2), m.group(3)
    insert_pos = len(body)
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"', body):
        if col_to_idx(cm.group(1)) > col_idx:
            insert_pos = cm.start()
            break
    body = body[:insert_pos] + new_cell + body[insert_pos:]
    return xml[:m.start()] + head + body + tail + xml[m.end():]


def _split_xfs(body):
    """Аккуратно разбивает cellXfs на отдельные <xf> (учёт самозакрытия и <alignment>)."""
    xfs = []
    for m in re.finditer(r'<xf\b', body):
        start = m.start()
        rest = body[start:]
        gt = rest.find(">")
        if gt > 0 and rest[gt - 1] == "/":
            end = start + gt + 1
        else:
            close = rest.find("</xf>")
            end = start + close + len("</xf>")
        xfs.append(body[start:end])
    return xfs


def _ensure_red_fill(styles_xml):
    """Гарантирует наличие красной заливки в палитре. Возвращает (xml, fill_id)."""
    m = re.search(r'(<fills count=")(\d+)(">)(.*?)(</fills>)', styles_xml, re.S)
    body = m.group(4)
    # уже есть красная (FFFF0000)?
    fills = re.findall(r"<fill>.*?</fill>", body, re.S)
    for i, f in enumerate(fills):
        if 'rgb="FFFF0000"' in f:
            return styles_xml, i
    red = '<fill><patternFill patternType="solid"><fgColor rgb="FFFF0000"/><bgColor indexed="64"/></patternFill></fill>'
    new_id = len(fills)
    styles_xml = (styles_xml[:m.start()] + m.group(1) + str(int(m.group(2)) + 1) + m.group(3)
                  + body + red + m.group(5) + styles_xml[m.end():])
    return styles_xml, new_id


def add_marker_style(styles_xml, fill_id, base_index=374):
    """Добавляет в cellXfs стиль = базовый(374) + заливка fill_id. Возвращает (xml, индекс)."""
    m = re.search(r'(<cellXfs count=")(\d+)(">)(.*?)(</cellXfs>)', styles_xml, re.S)
    if not m:
        return styles_xml, None
    body = m.group(4)
    xfs = _split_xfs(body)
    base = xfs[base_index] if base_index < len(xfs) else '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
    new = base
    if 'fillId="' in new:
        new = re.sub(r'fillId="\d+"', f'fillId="{fill_id}"', new, count=1)
    else:
        new = new.replace("<xf", f'<xf fillId="{fill_id}"', 1)
    if "applyFill=" in new:
        new = re.sub(r'applyFill="\d"', 'applyFill="1"', new, count=1)
    else:
        new = new.replace("<xf", '<xf applyFill="1"', 1)
    new_index = len(xfs)
    styles_xml = (styles_xml[:m.start()] + m.group(1) + str(int(m.group(2)) + 1) + m.group(3)
                  + body + new + m.group(5) + styles_xml[m.end():])
    return styles_xml, new_index


def ensure_full_calc(wb):
    if "fullCalcOnLoad" in wb:
        return wb
    m = re.search(r"<calcPr\b[^>]*?/>", wb)
    if m:
        return wb.replace(m.group(0), m.group(0)[:-2] + ' fullCalcOnLoad="1"/>', 1)
    return re.sub(r"(</sheets>)", r'\1<calcPr fullCalcOnLoad="1"/>', wb, count=1)


def read_svodnaya_names(xlsx_bytes):
    """Читает названия материалов из листа «Сводная» (столбец C, кэш-значения формул)."""
    entries, _ = read_svodnaya_full(xlsx_bytes)
    return [nm for _, nm, _ in entries]


def read_svodnaya_full(xlsx_bytes):
    """Возвращает ([(row, name, price)], [free_rows]) из листа «Сводная» (C=имя, G=цена)."""
    try:
        z = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
        s6 = z.read("xl/worksheets/sheet6.xml").decode("utf-8")
    except Exception:
        return [], []

    def cv(ref):
        m = re.search(r'<c r="' + ref + r'"[^>]*?>(?:<f>[^<]*</f>)?<v>([^<]*)</v>', s6)
        return m.group(1) if m else None

    entries, free = [], []
    for r in range(8, 621):
        c = cv(f"C{r}")
        if c and any(ch.isalpha() for ch in c):
            g = cv(f"G{r}")
            try:
                price = float(g)
            except (TypeError, ValueError):
                price = None
            entries.append((r, c, price))
        elif c is None:
            free.append(r)
    return entries, free


def make_real_name(real_size, analog_name, typ):
    """Подставляет реальный размер в формат имени аналога (для дозаписи в «Сводную»)."""
    real = int(real_size) if float(real_size).is_integer() else real_size
    if typ == "БАЛКА":
        return re.sub(r"(,\s*)\d+", lambda m: m.group(1) + str(real), analog_name, count=1)
    if typ == "ТРУБА_ПРОФ":
        return re.sub(r"\d+х\d+", f"{real}х{real}", analog_name, count=1)
    return re.sub(r"\d+", str(real), analog_name, count=1)


def _parse_profile(name):
    """(тип, [размеры], марка, подтип) из названия — нашего или из «Сводной»."""
    s = name.lower().replace("ё", "е")
    grade = "09Г2С" if "09г2с" in s else ("ст3" if ("ст 3" in s or "ст3" in s) else None)

    def nums(txt):
        txt = re.sub(r"(\d),(\d)", r"\1.\2", txt)  # 4,0 -> 4.0
        return [float(x) for x in re.findall(r"\d+(?:\.\d+)?", txt)]

    if "швеллер" in s:
        return ("ШВЕЛЛЕР", nums(s.split(",")[0].replace("швеллер", ""))[:1], grade, None)
    if "уголок" in s:
        a = nums(s.split("уголок")[1].split(",")[0])
        sub = None
        if len(a) >= 3:
            if a[0] != a[1]:
                sub = "неравн"            # неравнополочный
            dims = [max(a[:-1]), a[-1]]   # бОльшая полка + толщина
        elif len(a) == 2:
            dims = [a[0], a[1]]
        else:
            dims = a[:1]
        return ("УГОЛОК", dims, grade, sub)
    if "лист" in s:
        body = s.split("лист")[1].split(",")[0].replace("ст 3", "").replace("t", "")
        return ("ЛИСТ", nums(body)[:1], grade, None)
    if "балка" in s or "двутавр" in s:
        after = s.split("балка")[-1] if "балка" in s else s.split("двутавр")[-1]
        m_our = re.search(r"(\d+)\s*([кбш])", after)  # наш формат "30к1"
        if m_our:
            return ("БАЛКА", [float(m_our.group(1))], grade, m_our.group(2).upper())
        m_sub = re.match(r"\s*([кбшм])\b", after)      # Сводная "к-1, 30"
        sub = m_sub.group(1).upper() if m_sub else None
        size = None
        for part in after.split(",")[1:]:
            d = re.findall(r"\d+", part)
            if d:
                size = float(d[0]); break
        return ("БАЛКА", [size] if size else [], grade, sub)
    if "труб" in s:
        is_prof = "проф" in s
        body = re.sub(r"(\d),(\d)", r"\1.\2", s).split(",")[0]  # отрезать ", 12м"
        n = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", body)]
        if is_prof:
            if len(n) >= 3:
                sides = sorted(n[:-1], reverse=True)[:2]
                dims = [sides[0], sides[1], n[-1]]   # [большая сторона, меньшая, толщина]
            elif len(n) == 2:
                dims = [n[0], n[0], n[1]]            # квадрат «250x6» = 250×250×6
            else:
                dims = n
            return ("ТРУБА_ПРОФ", dims, grade, None)
        return ("ТРУБА_КРУГ", n, grade, None)
    if "катанка" in s:
        return ("КАТАНКА", nums(s.split(",")[0])[:1], grade, None)
    if "круг" in s or "арматур" in s:
        # «Круг 20», «Круг 20, 12м», «Арматура А-III, 20 мм» -> диаметр
        body = s.replace("круг", " ").replace("арматура", " ")
        n = nums(body)
        # выкинуть «12» от «12м» (длина хлыста), если затесалась последней
        return ("КРУГ", n[:1], grade, None)
    return ("?", [], grade, None)


def build_svodnaya_index(names):
    index = {}
    for nm in names:
        p = _parse_profile(nm)
        index.setdefault(p[0], []).append((p, nm))
    return index


def _grade_class(g):
    """Класс марки для выбора варианта в «Сводной»: '09Г2С' (низколегир.) | 'ст3' (углерод.) | None."""
    g = (g or "").upper().replace(" ", "")
    if any(k in g for k in ("09Г2С", "10ХСНД", "С345", "С375", "С390", "С440")):
        return "09Г2С"
    if any(k in g for k in ("С235", "С245", "С255", "С285", "СТ3", "ВСТ3")):
        return "ст3"
    return None


def resolve_to_svodnaya(name, grade, index):
    """Возвращает (имя_из_Сводной, способ). способ: 'точно' | 'аналог' | 'нет'."""
    p = _parse_profile(name)
    typ, dims, _, sub = p
    if typ == "УГОЛОК" and "гнут" in name.lower():
        return name, "гнутый"
    cands = index.get(typ, [])
    if not cands:
        return name, "нет"
    gcls = _grade_class(grade)
    if typ == "ЛИСТ":
        want = gcls                       # марка листа берётся с чертежа: С245→ст3, С345→09Г2С
    elif typ == "УГОЛОК" and gcls == "09Г2С":
        want = "09Г2С"
    else:
        want = None

    def gpref(lst):
        if want:
            g = [x for x in lst if x[0][2] == want]
            if g:
                return g
        ng = [x for x in lst if x[0][2] in (None, "ст3")]
        return ng or lst

    def same_sub(pp):
        return not (typ == "БАЛКА" and sub and pp[3] and pp[3] != sub)

    exact = [(pp, nm) for pp, nm in cands if pp[1] == dims and same_sub(pp)]
    if exact:
        return gpref(exact)[0][1], "точно"

    # ближайший больший
    if typ == "ТРУБА_ПРОФ" and len(dims) >= 3:
        big, small, wall = dims[0], dims[1], dims[2]
        # 1) то же сечение (обе стороны), толщина >= нужной — ближайшая большая толщина
        same_sec = [(pp[1][2], nm) for pp, nm in cands
                    if len(pp[1]) >= 3 and pp[1][0] == big and pp[1][1] == small and pp[1][2] >= wall]
        if same_sec:
            return min(same_sec)[1], "аналог"
        # 2) обе стороны >= нужных и толщина >= нужной — ближайший по площади сечения
        bigger = [(pp[1][0] * pp[1][1], nm) for pp, nm in cands
                  if len(pp[1]) >= 3 and pp[1][0] >= big and pp[1][1] >= small and pp[1][2] >= wall]
        if bigger:
            return min(bigger)[1], "аналог"
        # 3) обе стороны >= нужных, толщина любая
        bigger2 = [(pp[1][0] * pp[1][1], nm) for pp, nm in cands
                   if len(pp[1]) >= 3 and pp[1][0] >= big and pp[1][1] >= small]
        if bigger2:
            return min(bigger2)[1], "аналог"
    elif typ == "УГОЛОК" and len(dims) >= 2:
        leg, th = dims[0], dims[1]
        legs = sorted({pp[1][0] for pp, nm in cands if pp[1]})
        bigger = [L for L in legs if L >= leg]
        if bigger:
            std_leg = bigger[0]
            same = gpref([(pp, nm) for pp, nm in cands
                          if pp[1] and pp[1][0] == std_leg and len(pp[1]) >= 2 and pp[1][1] == th])
            if same:
                return same[0][1], "аналог"
        # нет стандартного с такой толщиной на ближайшей полке -> гнутый
        return name, "гнутый"
    else:
        big = sorted([(pp[1][0], pp, nm) for pp, nm in cands if pp[1] and dims and pp[1][0] >= dims[0] and same_sub(pp)], key=lambda x: x[0])
        big = gpref([(pp, nm) for _, pp, nm in big])
        if big:
            return big[0][1], "аналог"
    return name, "нет"


def match_sortament(name):
    # Канонизация имени профильной трубы к единому виду «Труба проф A*B*C»
    # (как в «Сводной»): слово «Труба проф», размеры через «*», квадрат из двух
    # чисел (250x6) разворачивается в три (250*250*6). Совпадение с ключом
    # справочника и ячейкой файла гарантируется одинаковым написанием везде.
    s = (name or "").strip()
    low = s.lower()
    if "труб" in low and "проф" in low:
        head = s.split(",")[0]
        tail = s[len(head):]                       # сохранить «, 12м» если было
        nums = re.findall(r"\d+(?:[.,]\d+)?", head)
        if len(nums) == 2:                         # квадрат «250x6» -> 250*250*6
            a, t = nums
            dims = [a, a, t]
        elif len(nums) >= 3:                       # прямоугольный 160x120x4
            dims = nums[:3]
        else:
            dims = None
        if dims:
            s = "Труба проф " + "*".join(dims) + tail
    return s


def _norm_key(name):
    # ключ для дедупа: убрать "ГОСТ ...", лишние пробелы, привести к нижнему регистру
    s = re.sub(r"\s*ГОСТ[\s\S]*$", "", name, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _resolve_one_material(raw, grade, svod_index, price_map, dopisat):
    """Разрешает один сортамент в (имя, цвет, коммент) по «Сводной». Выносено из
    build_raschet_cells, используется и одиночным, и мульти-заполнением."""
    uncertain = False
    typ = _parse_profile(raw)[0]
    dim0 = (_parse_profile(raw)[1] or [None])[0]
    name, mark, comment = raw, None, ""
    if svod_index:
        resolved, how = resolve_to_svodnaya(raw, grade, svod_index)
        if how == "точно":
            name, mark = resolved, None
            if typ == "УГОЛОК" and _parse_profile(raw)[3] == "неравн":
                mark = "yellow"
                comment = f"Неравнополочный уголок сведён к равнополочному {resolved}. Проверить."
        elif how == "гнутый":
            a = [x.replace(",", ".") for x in re.findall(r"\d+(?:[.,]\d+)?", raw)]
            def _f(v):
                return str(int(float(v))) if float(v).is_integer() else v
            body = "*".join(_f(x) for x in a) if a else ""
            real = f"Уголок {body}".strip()
            if "гнут" not in real.lower():
                real += " гнутый"
            name, mark = real, "red"
            comment = "Гнутый уголок (нет стандартного с такой толщиной полки). Запросить цену у менеджера."
        elif how == "аналог":
            if typ == "ЛИСТ":
                name, mark = resolved, "yellow"
                comment = f"Лист заменён на ближайший: {resolved}."
            elif typ in ("ТРУБА_ПРОФ", "ТРУБА_КРУГ"):
                name, mark = raw, "orange"
                dopisat.append((raw, price_map.get(resolved)))
                comment = f"Нет в базе. Цена от аналога {resolved}. Проверить."
            elif typ == "УГОЛОК":
                name, mark = resolved, "yellow"
                comment = f"Уголок заменён на ближайший стандартный: {resolved}. Цена от него."
            else:
                real = make_real_name(dim0, resolved, typ) if dim0 else raw
                name, mark = real, "orange"
                dopisat.append((real, price_map.get(resolved)))
                comment = f"Нет в базе. Цена от аналога {resolved}. Проверить."
        else:
            name, mark = raw, "red"
            comment = "Не найдено в базе. Проверить и добавить цену."
    return name, mark, comment


def build_raschet_cells_multi(products, svod_index=None, price_map=None):
    """Заполнение НЕСКОЛЬКИХ изделий: каждое — своя строка (6,7,8…),
    материалы — общие столбцы (N,O,P…) на весь заказ.
    products = [{"mark": str, "qty": int, "positions": [..]}, ...]
    Возвращает (cells, notes, dopisat, summary, per_product) как build_raschet_cells,
    плюс per_product = [{"row":r,"mark":..,"mass":..}] для самопроверки.
    """
    price_map = price_map or {}
    cells = {}
    notes, dopisat = [], []
    summary = {"yellow": 0, "orange": 0, "red": 0}
    per_product = []

    # 1) общая карта материал -> столбец (по порядку появления во всех изделиях)
    col_of = {}        # _norm_key(имя) -> столбец
    mat_meta = {}      # _norm_key -> (имя, mark, comment)
    next_col = 0

    ROW0 = 6
    for pi, prod in enumerate(products):
        row = ROW0 + pi
        if row - ROW0 >= 20:        # ограничение: до 20 изделий
            break
        mark_name = prod.get("mark") or prod.get("name") or f"Изделие {pi+1}"
        qty = prod.get("qty") or 1
        cells[f"C{row}"] = (mark_name, "s", None)
        cells[f"D{row}"] = (prod.get("drawing") or "", "s", None)
        cells[f"I{row}"] = ("шт", "s", None)
        cells[f"J{row}"] = (qty, "n", None)

        # масса каждого материала этого изделия
        per_mat = {}    # столбец -> масса
        for p in prod.get("positions", []):
            raw = match_sortament(p.get("sortament") or p.get("name") or "")
            if not raw:
                continue
            try:
                mass = float(p.get("mass_kg") or p.get("mass") or 0)
            except (TypeError, ValueError):
                mass = 0.0
            if mass <= 0:
                continue
            grade = str(p.get("grade") or "")
            name, mk, comment = _resolve_one_material(raw, grade, svod_index, price_map, dopisat)
            key = _norm_key(name)
            if key not in col_of:
                if next_col >= len(STEEL_COLUMNS):
                    continue   # столбцы кончились — пропускаем (предупреждение в summary)
                col_of[key] = STEEL_COLUMNS[next_col]
                mat_meta[key] = (name, mk, comment)
                next_col += 1
                if mk:
                    notes.append({"исходное": raw, "имя": name, "пометка": mk, "коммент": comment})
                    if mk in summary:
                        summary[mk] += 1
            col = col_of[key]
            per_mat[col] = per_mat.get(col, 0.0) + mass

        # пишем массы материалов в строку этого изделия
        total = 0.0
        for col, m in per_mat.items():
            cells[f"{col}{row}"] = (round(m, 2), "n", None)
            total += m
        cells[f"K{row}"] = (round(total, 2), "n", None)
        per_product.append({"row": row, "mark": mark_name, "mass": round(total, 2),
                            "qty": qty, "control": prod.get("control_total_kg"),
                            "weld_pct": prod.get("weld_pct")})

    # шапка материалов (строка 5) + комментарии (строка 1)
    for key, col in col_of.items():
        name, mk, comment = mat_meta[key]
        cells[f"{col}5"] = (name, "s", mk)
        if comment:
            cells[f"{col}1"] = (comment, "s", None)

    return cells, notes, dopisat, summary, per_product


def build_raschet_cells(positions, drawing_name, svod_index=None, price_map=None):
    """
    Возвращает (cells, notes, dopisat, summary).
    cells[ref] = (значение, тип, mark)  где mark = None | "yellow" | "orange" | "red"
    comment-ячейки (строка 7) тоже в cells (без цвета).
    dopisat = [(имя, цена)] — что дописать в «Сводную» (оранжевые).
    summary = {"yellow":n,"orange":n,"red":n}
    """
    price_map = price_map or {}
    cells = {
        "C6": ("Металлоконструкции", "s", None),
        "D6": (drawing_name or "", "s", None),
        "I6": ("шт", "s", None),
        "J6": (1, "n", None),
    }
    agg, info, order, notes, dopisat = {}, {}, [], [], []
    summary = {"yellow": 0, "orange": 0, "red": 0}

    for p in positions:
        raw = match_sortament(p.get("sortament") or p.get("name") or "")
        if not raw:
            continue
        try:
            mass = float(p.get("mass_kg") or p.get("mass") or 0)
        except (TypeError, ValueError):
            mass = 0.0
        if mass <= 0:
            continue
        grade = str(p.get("grade") or "")
        uncertain = bool(p.get("uncertain"))
        typ = _parse_profile(raw)[0]
        dim0 = (_parse_profile(raw)[1] or [None])[0]

        name, mark, comment = raw, None, ""
        if uncertain:
            name, mark = raw, "red"
            comment = "ВНИМАНИЕ: неуверенно прочитано с чертежа. Проверить вручную."
        elif svod_index:
            resolved, how = resolve_to_svodnaya(raw, grade, svod_index)
            if how == "точно":
                name, mark = resolved, None
                if typ == "УГОЛОК" and _parse_profile(raw)[3] == "неравн":
                    mark = "yellow"
                    comment = f"Неравнополочный уголок сведён к равнополочному {resolved}. Проверить."
            elif how == "гнутый":
                # гнутый уголок: стандартного с такой толщиной нет — оставляем
                # реальную геометрию, помечаем красным, цену ставит менеджер
                a = [x.replace(",", ".") for x in re.findall(r"\d+(?:[.,]\d+)?", raw)]
                def _f(v):
                    return str(int(float(v))) if float(v).is_integer() else v
                body = "*".join(_f(x) for x in a) if a else ""
                real = f"Уголок {body}".strip()
                if "гнут" not in real.lower():
                    real += " гнутый"
                name, mark = real, "red"
                comment = "Гнутый уголок (нет стандартного с такой толщиной полки). Запросить цену у менеджера."
            elif how == "аналог":
                if typ == "ЛИСТ":
                    name, mark = resolved, "yellow"
                    comment = f"Лист заменён на ближайший: {resolved}."
                elif typ in ("ТРУБА_ПРОФ", "ТРУБА_КРУГ"):
                    name, mark = raw, "orange"
                    price = price_map.get(resolved)
                    dopisat.append((raw, price))
                    comment = f"Нет в базе. Цена от аналога {resolved}. Проверить."
                elif typ == "УГОЛОК":
                    # полка округлена до ближайшей стандартной; resolved РЕАЛЬНО есть
                    # в «Сводной» со своей ценой -> жёлтый, дозапись не нужна
                    name, mark = resolved, "yellow"
                    comment = f"Уголок заменён на ближайший стандартный: {resolved}. Цена от него."
                else:
                    real = make_real_name(dim0, resolved, typ) if dim0 else raw
                    name, mark = real, "orange"
                    price = price_map.get(resolved)
                    dopisat.append((real, price))
                    comment = f"Нет в базе. Цена от аналога {resolved}. Проверить."
            else:  # не нашли совсем
                name, mark = raw, "red"
                comment = "Не найдено в базе. Проверить и добавить цену."
        else:
            name, mark = raw, None

        key = _norm_key(name)
        if key not in agg:
            agg[key] = 0.0
            info[key] = (name, mark, comment)
            order.append(key)
        agg[key] = max(agg[key], mass)
        if mark in ("аналог",):
            pass
        if mark:
            notes.append({"исходное": raw, "имя": name, "пометка": mark, "коммент": comment})

    total = 0.0
    for i, key in enumerate(order):
        if i >= len(STEEL_COLUMNS):
            break
        c = STEEL_COLUMNS[i]
        name, mark, comment = info[key]
        cells[f"{c}5"] = (name, "s", mark)             # имя (с цветом)
        cells[f"{c}6"] = (round(agg[key], 2), "n", None)  # масса
        if comment:
            cells[f"{c}1"] = (comment, "s", None)      # комментарий сверху, над названием
        if mark in summary:
            summary[mark] += 1
        total += agg[key]
    cells["K6"] = (round(total, 2), "n", None)
    return cells, notes, dopisat, summary


def apply_catalog_to_template(xlsx_bytes, catalog):
    """Применяет глобальные цены справочника к «Сводной» шаблона ДО заполнения.
    Цена есть в «Сводной» -> перезаписываем; нет -> добавляем в свободную строку.
    Структура файла сохраняется (правка XML); при ошибке возвращаем исходный шаблон."""
    if not catalog:
        return xlsx_bytes
    try:
        entries, free_rows = read_svodnaya_full(xlsx_bytes)
        zin = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
        parts = {n: zin.read(n) for n in zin.namelist()}
        name_to_row = {nm: r for r, nm, _ in entries}
        free_iter = iter(free_rows)
        s6 = parts["xl/worksheets/sheet6.xml"].decode("utf-8")
        changed = False
        for nm, price in catalog.items():
            if price is None:
                continue
            row = name_to_row.get(nm)
            if row is None:
                row = next(free_iter, None)
                if row is None:
                    continue
                s6 = set_cell(s6, f"C{row}", nm, "s")
                name_to_row[nm] = row
            s6 = set_cell(s6, f"G{row}", price, "n")
            changed = True
        if not changed:
            return xlsx_bytes
        parts["xl/worksheets/sheet6.xml"] = s6.encode("utf-8")
        out = io.BytesIO()
        zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
        for zi in zin.infolist():
            zout.writestr(zi, parts[zi.filename])
        zout.close()
        return out.getvalue()
    except Exception as e:
        log.error("[catalog] не удалось применить справочник к шаблону: %s", e)
        return xlsx_bytes


def fill_template(xlsx_bytes, positions, drawing_name, overrides=None, catalog=None, pricelist=None, products=None, diag_text=None):
    zin = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    parts = {n: zin.read(n) for n in zin.namelist()}

    entries, free_rows = read_svodnaya_full(xlsx_bytes)
    price_map = {nm: pr for _, nm, pr in entries}
    svod_index = build_svodnaya_index([nm for _, nm, _ in entries])
    per_product = None
    if products:
        cells, notes, dopisat, summary, per_product = build_raschet_cells_multi(
            products, svod_index, price_map)
    else:
        cells, notes, dopisat, summary = build_raschet_cells(
            positions, drawing_name, svod_index, price_map)

    # служебная диагностика в A1 (пустая, безопасная ячейка) — версия/листы
    if diag_text:
        cells["A1"] = (diag_text, "s", None)

    # стили цветов: для каждой помечаемой ячейки клонируем ЕЁ СОБСТВЕННЫЙ стиль
    # (шрифт, перенос, рамка) и только меняем заливку — формат ячейки сохраняется
    styles = parts["xl/styles.xml"].decode("utf-8")
    s1 = parts["xl/worksheets/sheet1.xml"].decode("utf-8")

    used = {m for _, _, m in cells.values() if m}
    fill_for = {}
    if "red" in used:
        styles, red_fill = _ensure_red_fill(styles)
    for mark in used:
        fill_for[mark] = red_fill if mark == "red" else {"yellow": 2, "orange": 6}[mark]

    def cur_style(ref):
        mm = re.search(r'<c r="' + ref + r'"([^>]*?)(?:/>|>)', s1)
        if mm:
            sm = re.search(r'\bs="(\d+)"', mm.group(1))
            if sm:
                return int(sm.group(1))
        return 0

    # лист «Расчет»
    marker_cache = {}   # (исходный стиль ячейки, заливка) -> индекс нового стиля
    for ref, (val, kind, mark) in cells.items():
        sid = None
        if mark:
            ckey = (cur_style(ref), fill_for[mark])
            if ckey not in marker_cache:
                styles, idx = add_marker_style(styles, ckey[1], base_index=ckey[0])
                marker_cache[ckey] = idx
            sid = marker_cache[ckey]
        s1 = set_cell(s1, ref, val, kind, sid)

    parts["xl/styles.xml"] = styles.encode("utf-8")
    parts["xl/worksheets/sheet1.xml"] = s1.encode("utf-8")

    # запись в «Сводную»: прайс металла (база) + новые материалы + справочник + цены заявки
    overrides = overrides or {}
    catalog = catalog or {}
    pricelist = pricelist or {}
    name_to_row = {nm: r for r, nm, _ in entries}
    free_iter = iter(free_rows)
    s6 = parts["xl/worksheets/sheet6.xml"].decode("utf-8")
    s6_changed = False

    def put_price(nm, price_per_kg):
        """Проставить цену материалу. Менеджер вводит цену ЗА КГ, а «Сводная»
        хранит цену ЗА ТОННУ (формула «Расчёта» делит на 1000) — поэтому ×1000.
        Есть в «Сводной» -> переписать G; нет -> добавить в свободную строку."""
        nonlocal s6, s6_changed
        if price_per_kg is None:
            return False
        price = price_per_kg * 1000          # за кг -> за тонну
        row = name_to_row.get(nm)
        if row is None:
            row = next(free_iter, None)
            if row is None:
                return False
            s6 = set_cell(s6, f"C{row}", nm, "s")
            name_to_row[nm] = row
        s6 = set_cell(s6, f"G{row}", price, "n")
        s6_changed = True
        return True

    # 0) базовые цены металла из загруженного прайса (за тонну, КАК ЕСТЬ — без ×1000),
    #    только по точному совпадению имени; перезаписываем формулу-ссылку [1]металл числом.
    #    Чего нет в прайсе (болты, спецбалки) — не трогаем, остаётся прежнее значение.
    pl_applied = 0
    for nm, price_ton in pricelist.items():
        row = name_to_row.get(nm)
        if row is not None and price_ton is not None:
            s6 = set_cell(s6, f"G{row}", price_ton, "n")
            s6_changed = True
            pl_applied += 1

    # 1) новые материалы (оранжевые/трубы), которых ещё нет в «Сводной»
    for real_name, price in dopisat:
        if real_name in name_to_row:
            continue
        row = next(free_iter, None)
        if row is None:
            break
        s6 = set_cell(s6, f"C{row}", real_name, "s")
        name_to_row[real_name] = row
        if price is not None:
            s6 = set_cell(s6, f"G{row}", price, "n")
        s6_changed = True

    # 2) справочник (на все заявки)
    for nm, price in catalog.items():
        put_price(nm, price)

    # 3) ручные цены менеджера по заявке — перекрывают справочник
    for nm, price in overrides.items():
        put_price(nm, price)

    if s6_changed:
        parts["xl/worksheets/sheet6.xml"] = s6.encode("utf-8")

    wb = parts["xl/workbook.xml"].decode("utf-8")
    parts["xl/workbook.xml"] = ensure_full_calc(wb).encode("utf-8")

    # Мы перезаписали формульные ячейки (K6, C6, N5/N6 …) значениями, а
    # xl/calcChain.xml всё ещё считает их формулами — из-за этого Excel при
    # открытии показывает «Ошибка в части содержимого… восстановить?».
    # Удаляем calcChain (и две ссылки на него) — Excel строит цепочку заново,
    # тем более что fullCalcOnLoad уже включён.
    parts.pop("xl/calcChain.xml", None)
    ct = parts["[Content_Types].xml"].decode("utf-8")
    ct = re.sub(r'<Override[^>]*calcChain\.xml[^>]*/>', "", ct)
    parts["[Content_Types].xml"] = ct.encode("utf-8")
    rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rels = re.sub(r'<Relationship[^>]*calcChain\.xml[^>]*/>', "", rels)
    parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    out = io.BytesIO()
    zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
    for zi in zin.infolist():
        if zi.filename not in parts:      # пропускаем удалённый calcChain
            continue
        zout.writestr(zi, parts[zi.filename])
    zout.close()
    return out.getvalue(), notes, summary, per_product


# ---------------------------------------------------------------------------
# Фоновая обработка
# ---------------------------------------------------------------------------
STEEL_DENSITY = 7850.0  # кг/м3


def _plate_to_product(plate, filename, request_items, phase):
    """Из размеров пластины (ТИП 4) считает массу заготовки и собирает изделие.
    Масса = длина×ширина×толщина×плотность (отверстия НЕ вычитаются)."""
    import os as _os
    L = _safe_float(plate.get("length_mm"))
    W = _safe_float(plate.get("width_mm"))
    S = _safe_float(plate.get("thickness_mm"))
    grade = (plate.get("grade") or "").strip()
    name = (plate.get("name") or "").strip() or \
        _os.path.splitext(_os.path.basename(filename))[0].replace("_", " ")
    holes = plate.get("holes") or []

    if L <= 0 or W <= 0 or S <= 0:
        phase("PLATE_BAD_DIMS", length_mm=L, width_mm=W, thickness_mm=S)
        return None
    mass = round(L / 1000.0 * W / 1000.0 * S / 1000.0 * STEEL_DENSITY, 2)
    sortament = f"Лист {int(S) if float(S).is_integer() else S}"
    phase("PLATE_CALC", det=name[:60], L=L, W=W, S=S, grade=grade,
          mass_kg=mass, sortament=sortament,
          holes=[{"d": h.get("d"), "count": h.get("count")} for h in holes])

    pos = [{"sortament": sortament, "grade": grade or "09Г2С", "mass_kg": mass, "qty": 1}]
    qty = 1
    if request_items:
        q, qstatus, _m = match_request_to_drawing(request_items, name)
        if qstatus == "green" and q:
            qty = q
    return {
        "mark": name, "drawing": "", "qty": qty, "positions": pos,
        "control_total_kg": mass, "weld_pct": None,
        "mass_kg": mass, "control_no_weld_kg": mass, "diff_pct": 0.0,
        "check_status": "CHECK_OK", "holes": holes,
    }


def _finalize_product(data, filename, request_items, phase, mark_default=None):
    """Из ответа модели (data с positions/mark/control) собирает product dict
    с самопроверкой массы. mark_default — имя изделия, если в штампе марки нет."""
    import os as _os
    positions = data.get("positions") or []
    # ТИП 2: суммируем по сортаменту в коде
    if len(positions) > len(_uniq_sortaments(positions)):
        positions = _aggregate_by_sortament(positions)
        phase("AGGREGATED", groups=len(positions))

    fname_mark = _os.path.splitext(_os.path.basename(filename))[0].replace("_", " ")
    mark = data.get("mark") or mark_default or data.get("drawing") or fname_mark

    # количество из заявки по марке (только зелёная связка)
    qty = 1
    if request_items:
        q, qstatus, matched = match_request_to_drawing(request_items, mark)
        phase("REQUEST_MATCH", drawing_mark=str(mark)[:60], qty=q, status=qstatus,
              matched_mark=(matched or {}).get("mark"))
        if qstatus == "green" and q:
            qty = q

    # самопроверка массы (с поправкой на сварку, вычет один раз)
    sum_mass = round(sum(
        float(p.get("mass_kg") or p.get("mass") or 0)
        for p in positions
        if _safe_float(p.get("mass_kg") or p.get("mass")) > 0), 2)
    control = _safe_float(data.get("control_total_kg")) or None
    weld_pct = _safe_float(data.get("weld_pct")) or None
    control_noweld = control
    if control and weld_pct:
        control_noweld = round(control / (1.0 + weld_pct / 100.0), 2)
    if control_noweld and control_noweld > 0:
        diff_pct = round(abs(sum_mass - control_noweld) / control_noweld * 100, 1)
        cstatus = "CHECK_OK" if diff_pct <= 1.0 else "CHECK_MISMATCH"
    else:
        diff_pct = None
        cstatus = "CHECK_NO_CONTROL"
    phase(cstatus, mark=str(mark)[:60], sum_mass_kg=sum_mass,
          control_no_weld_kg=control_noweld, diff_pct=diff_pct)

    return {
        "mark": mark, "drawing": data.get("drawing") or "",
        "qty": qty, "positions": positions,
        "control_total_kg": control, "weld_pct": weld_pct,
        "mass_kg": sum_mass, "control_no_weld_kg": control_noweld,
        "diff_pct": diff_pct, "check_status": cstatus,
    }


def _split_sheet_into_products(data, filename, request_items, phase, sheet, seen_all):
    """Делит прочитанный лист на ОТДЕЛЬНЫЕ изделия по полю "izdelie".
    Если разметки нет — одно изделие «Лист N» (как раньше). Каждое изделие
    получает свой контроль из заявленных масс "izdelia"."""
    positions = data.get("positions") or []
    izd_mass = {}
    for it in (data.get("izdelia") or []):
        nm = str(it.get("name") or "").strip()
        m = _safe_float(it.get("mass_kg"))
        if nm and m > 0:
            izd_mass[nm] = m

    # группируем позиции по изделию (сохраняя порядок появления)
    groups, order = {}, []
    for p in positions:
        key = str(p.get("izdelie") or "").strip() or f"Лист {sheet}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(p)

    out = []
    for name in order:
        sub = {
            "positions": groups[name],
            "mark": name,
            "drawing": data.get("drawing") or "",
            "weld_pct": data.get("weld_pct"),
            "control_total_kg": izd_mass.get(name),
        }
        prod = _finalize_product(sub, filename, request_items, phase, mark_default=name)
        prod["sheet"] = sheet
        prod["sheets_seen"] = seen_all
        out.append(prod)
    return out


def recognize_one_drawing(draw_bytes, filename, request_items, phase, target_sheets=None):
    """Распознаёт чертёж -> СПИСОК product dict (одно изделие = одна строка Расчёта).
    Если указаны target_sheets (напр. [28, 30]) — каждый лист читается ОТДЕЛЬНЫМ
    проходом и даёт своё изделие (своя строка). Иначе — один общий проход."""
    import os as _os
    images = drawing_to_images(draw_bytes, filename)
    base = _os.path.basename(filename)

    if target_sheets:
        products = []
        # 1) карта: какой лист на какой странице (по штампу), дешёвый проход
        try:
            pmap = map_pages_to_sheets(images)
        except Exception as e:
            pmap = {}
            phase("PAGEMAP_ERROR", error=str(e)[:200])
        seen_all = sorted(pmap.keys()) if pmap else []
        phase("PAGE_MAP", map={k: v + 1 for k, v in pmap.items()}, sheets_seen=seen_all)

        for sheet in target_sheets:
            key = str(sheet).strip().lower()
            t = time.time()
            if key in pmap:
                pidx = pmap[key]
                long_pt = _page_long_side_pt(draw_bytes, filename, pidx)
                is_large = bool(long_pt and long_pt > LARGE_FMT_PT)
                if is_large:
                    # большой формат — сразу тяжёлый рендер с тайлами
                    imgs_read = render_page_tiles(draw_bytes, filename, pidx)
                    mode = f"tiles(large,{len(imgs_read)})"
                    data = extract_positions(imgs_read, target_sheets=[sheet], preselected=True)
                else:
                    # обычный формат — сначала лёгкий проход (дёшево)
                    imgs_read = render_page_single(draw_bytes, filename, pidx)
                    data = extract_positions(imgs_read, target_sheets=[sheet], preselected=True)
                    mode = "single"
                    if len(data.get("positions") or []) == 0:
                        # таблица не прочиталась — эскалация в тяжёлый рендер
                        imgs_read = render_page_tiles(draw_bytes, filename, pidx)
                        mode = f"tiles(escalated,{len(imgs_read)})"
                        data = extract_positions(imgs_read, target_sheets=[sheet], preselected=True)
                phase("CLAUDE_CALL", file=base, model=CLAUDE_MODEL, sheet=sheet,
                      page=pidx + 1, long_pt=round(long_pt or 0), mode=mode)
            else:
                # штамп этого листа не найден в карте — пробуем по всем страницам (запас)
                phase("CLAUDE_CALL", file=base, model=CLAUDE_MODEL, sheet=sheet,
                      page=None, mode="all-pages", note="лист не в карте штампов")
                data = extract_positions(images, target_sheets=[sheet])
            n = len(data.get("positions") or [])
            phase("CLAUDE_DONE", ms=int((time.time() - t) * 1000), positions=n,
                  sheet=sheet, sheets_seen=seen_all)
            if n == 0:
                phase("SHEET_EMPTY", sheet=sheet, sheets_seen=seen_all,
                      in_map=(key in pmap))
                continue
            sheet_prods = _split_sheet_into_products(
                data, filename, request_items, phase, sheet, seen_all)
            phase("SHEET_SPLIT", sheet=sheet, izdelia=[p["mark"] for p in sheet_prods])
            products.extend(sheet_prods)
        if not products:
            return [{
                "mark": "ЛИСТ НЕ НАЙДЕН", "drawing": "", "qty": 1, "positions": [],
                "control_total_kg": None, "weld_pct": None, "mass_kg": 0.0,
                "control_no_weld_kg": None, "diff_pct": None,
                "check_status": "CHECK_NO_CONTROL",
                "sheets_seen": seen_all, "empty_sheets": [str(s) for s in target_sheets],
            }]
        return products

    phase("CLAUDE_CALL", pages=len(images), file=base, model=CLAUDE_MODEL)
    t = time.time()
    data = extract_positions(images)
    phase("CLAUDE_DONE", ms=int((time.time() - t) * 1000),
          positions=len(data.get("positions") or []))

    # ТИП 4 — деталь-пластина (эскиз без таблицы): масса по геометрии
    plates = data.get("plates")
    if not plates and data.get("plate"):
        plates = [data["plate"]]
    if plates:
        out = []
        for pl in plates:
            prod = _plate_to_product(pl, filename, request_items, phase)
            if prod:
                out.append(prod)
        if out:
            return out

    return [_finalize_product(data, filename, request_items, phase)]


def _safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def process_order(order_id):
    order = fetch_order(order_id)
    entries = order.get(LOG_COL) or []
    t0 = time.time()

    def phase(name, **extra):
        entries.append({"ts": now_iso(), "step": name, **extra})
        log.info("[worker] %s %s", name, extra or "")
        db_update(order_id, {LOG_COL: entries})

    try:
        phase("LOAD_TEMPLATE")
        tpl = get_template_bytes()
        try:
            catalog = load_catalog_overrides()
            phase("CATALOG_LOADED", count=len(catalog), names=list(catalog.keys())[:20])
        except Exception as e:
            catalog = {}
            phase("CATALOG_ERROR", error=str(e)[:300])

        try:
            pricelist = load_metall_pricelist()
            phase("PRICELIST_LOADED", count=len(pricelist))
        except Exception as e:
            pricelist = {}
            phase("PRICELIST_ERROR", error=str(e)[:300])

        # --- Заявка: текст из формы (приоритет) или файл (PDF/картинки/Word/Excel) ---
        request_items = []
        req_text = (order.get(REQUEST_TEXT_COL) or "").strip()
        # номера листов из команды менеджера («посчитай лист 28, 30») — читаем металл только с них
        target_sheets = _parse_sheet_numbers(req_text)
        if target_sheets:
            phase("SHEET_FILTER", sheets=target_sheets)
        if req_text:
            try:
                request_items = extract_request_items_from_text(req_text).get("items", [])
                phase("REQUEST_PARSED", count=len(request_items), source="text")
            except Exception as e:
                phase("REQUEST_ERROR", error=str(e)[:300])
        else:
            req_data = download_request(order)
            if req_data:
                try:
                    request_items = read_request_items(
                        req_data, order.get(REQUEST_PATH_COL, "")).get("items", [])
                    phase("REQUEST_PARSED", count=len(request_items), source="file")
                except Exception as e:
                    phase("REQUEST_ERROR", error=str(e)[:300])

        # --- Список чертежей: из order_drawings, иначе одиночный drawing_url ---
        draw_rows = load_order_drawings(order_id)
        if draw_rows:
            phase("DRAWINGS_FOUND", count=len(draw_rows))
            sources = [(r.get("drawing_url"), r) for r in draw_rows if r.get("drawing_url")]
        else:
            phase("DRAWINGS_FOUND", count=1, mode="single")
            sources = [(order.get(DRAWING_PATH_COL), None)]

        products = []
        for path, drow in sources:
            try:
                phase("LOAD_DRAWING", file=path)
                db = download_drawing_by_path(path) if drow else download_drawing(order)
                prods = recognize_one_drawing(db, path or "", request_items, phase,
                                              target_sheets=target_sheets)
                products.extend(prods)
                # статус -> в order_drawings (если запись одна и изделие одно)
                if drow and len(prods) == 1:
                    prod = prods[0]
                    update_order_drawing(drow["id"], {
                        "mark": prod["mark"], "qty": prod["qty"],
                        "mass_kg": prod["mass_kg"], "control_kg": prod["control_no_weld_kg"],
                        "diff_pct": prod["diff_pct"], "check_status": prod["check_status"],
                    })
            except Exception as e:
                phase("DRAWING_ERROR", file=path, error=str(e)[:300])

        if not products:
            raise RuntimeError("не удалось распознать ни одного чертежа")

        # общий статус проверки: худший по изделиям
        statuses = [p["check_status"] for p in products]
        needs_review = any(s != "CHECK_OK" for s in statuses)
        overall = ("CHECK_MISMATCH" if "CHECK_MISMATCH" in statuses
                   else ("CHECK_NO_CONTROL" if "CHECK_NO_CONTROL" in statuses else "CHECK_OK"))
        phase("MASS_CHECK_OVERALL", status=overall, products=len(products),
              total_mass=round(sum(p["mass_kg"] for p in products), 2))

        phase("FILL_EXCEL")
        overrides = load_overrides(order_id)
        # диагностика в A1: версия / модель / какие листы просили / что увидели в штампах
        seen = next((p.get("sheets_seen") for p in products if p.get("sheets_seen")), [])
        diag_text = (f"{WORKER_VERSION} | model={CLAUDE_MODEL} | "
                     f"просили листы={target_sheets or '—'} | "
                     f"в штампах={seen or '—'} | изделий={len(products)}")
        filled, notes, summary, per_product = fill_template(
            tpl, [], "", overrides, catalog, pricelist, products=products, diag_text=diag_text)
        phase("DIAG_A1", text=diag_text)
        if notes:
            phase("SVODNAYA_ANALOGS", count=len(notes), items=notes)
        review_total = summary["yellow"] + summary["orange"] + summary["red"]
        phase("REVIEW_SUMMARY", **summary, total=review_total)

        phase("SAVE", bytes=len(filled))
        path = f"{order_id}.xlsx"
        upload_result(path, filled)
        result_fields = {RESULT_PATH_COL: path, STATUS_COL: STATUS_DONE}

        total_mass = round(sum(p["mass_kg"] for p in products), 2)
        worst = next((p for p in products if p["check_status"] == "CHECK_MISMATCH"),
                     next((p for p in products if p["check_status"] == "CHECK_NO_CONTROL"),
                          products[0]))
        # сводка на проверку для дашборда (поля могут отсутствовать в схеме — пишем мягко)
        try:
            db_update(order_id, {
                **result_fields,
                "review_total": review_total,
                "review_yellow": summary["yellow"],
                "review_orange": summary["orange"],
                "review_red": summary["red"],
                # контроль массы по заказу: суммарная масса + статус (худший по изделиям)
                "mass_sum_kg": total_mass,
                "mass_control_kg": worst.get("control_no_weld_kg"),
                "mass_diff_pct": worst.get("diff_pct"),
                "mass_check_status": overall,
                "weld_pct": worst.get("weld_pct"),
            })
        except Exception:
            db_update(order_id, result_fields)  # если полей нет в схеме — хотя бы статус

        phase("DONE", ms=int((time.time() - t0) * 1000),
              needs_review=needs_review, products=len(products), total_mass=total_mass)

        log_run(
            drawing_name    = ", ".join(p["mark"] for p in products)[:200],
            request_id      = order_id,
            check_status    = "OK" if overall == "CHECK_OK"
                              else ("MISMATCH" if overall == "CHECK_MISMATCH" else "NO_CONTROL"),
            positions_count = sum(len(p["positions"]) for p in products),
            sum_mass        = total_mass,
            expected_total  = worst.get("control_no_weld_kg"),
            review          = {"total":  review_total,
                               "yellow": summary["yellow"],
                               "orange": summary["orange"],
                               "red":    summary["red"]},
        )
    except Exception as e:
        tb = traceback.format_exc()
        log.error("[worker] FATAL %s", tb)
        entries.append({"ts": now_iso(), "step": "ERROR", "message": str(e), "traceback": tb[-1500:]})
        db_update(order_id, {LOG_COL: entries, STATUS_COL: "error", ERROR_MSG_COL: str(e)})


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
class ProcessReq(BaseModel):
    order_id: str


@app.get("/health")
def health():
    return {"ok": True, "version": WORKER_VERSION}


@app.post("/process")
def process(req: ProcessReq, background: BackgroundTasks,
            x_worker_secret: str = Header(default="", alias="X-Worker-Secret")):
    if not WORKER_SECRET or x_worker_secret != WORKER_SECRET:
        raise HTTPException(status_code=401, detail="bad secret")
    try:
        order = fetch_order(req.order_id)
    except Exception:
        order = None
    if not order:
        raise HTTPException(status_code=404, detail="order not found")

    entries = order.get(LOG_COL) or []
    entries.append({"ts": now_iso(), "step": "WORKER_START", "orderId": req.order_id})
    db_update(req.order_id, {STATUS_COL: "processing", LOG_COL: entries})

    background.add_task(process_order, req.order_id)
    return JSONResponse(status_code=202, content={"accepted": True})


@app.get("/result/{order_id}")
def get_result(order_id: str):
    """Прямое скачивание готового файла: отдаёт xlsx из бакета results.
    Открой в браузере: https://<worker>/result/<order_id>
    """
    from fastapi.responses import Response
    try:
        order = fetch_order(order_id)
    except Exception:
        order = None
    if not order:
        raise HTTPException(status_code=404, detail="order not found")
    path = order.get(RESULT_PATH_COL)
    if not path:
        raise HTTPException(status_code=404, detail="result not ready")
    path = path.lstrip("/").replace("results/", "", 1) if path.startswith("results/") else path.lstrip("/")
    try:
        data = sb.storage.from_(RESULTS_BUCKET).download(path)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"file not found: {e}")
    fname = f"{order.get('order_number') or order_id}.xlsx"
    return Response(
        content=data,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
